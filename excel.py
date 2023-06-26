import openpyxl as Excel
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelWriter:

	def __init__(self, az_services):
		self.wb = Excel.Workbook()
		self.az_services = az_services

		# Get workbook active sheet from the active attribute.
		sheet = self.wb.active
		# The first sheet will be the data summary sheet.
		sheet.title = 'Summary'

	def _auto_size_columns(self):
		for sheet_name in self.wb.sheetnames:
			for column_cells in self.wb[sheet_name].columns:
				new_column_width = 0
				for cell in column_cells:
					if cell.data_type == 's':
						lines = cell.value.splitlines()
						for line in lines:
							new_column_width = max(len(str(line)), new_column_width)
					elif cell.data_type == 'n':
						new_column_width = max(len(str(cell.value)), new_column_width)
					else:
						# This cell is probably a formula.
						# Therefore, use a larger number than we'll ever have
						# as the max length.
						new_column_width = max(len(str('99.99%')), new_column_width)

				new_column_letter = (get_column_letter(column_cells[0].column))
				if new_column_width > 0:
					self.wb[sheet_name].column_dimensions[new_column_letter].width = new_column_width * 1.23

	def save(self, file_name):
		self._auto_size_columns()
		self.wb.save(file_name)

	def write_data(self, article_excludes):
		self._write_sheet_azure_services()
		self._write_sheet_terraform_resources()
		self._write_sheet_excluded_articles(article_excludes)
		self._write_sheet_summary()

	def _write_sheet_summary(self):
		sheet = self.wb['Summary']

		# Write summary data.
		data = [
			['Total Azure services', '=COUNTA(AzureServicesArticles[Azure service])',],
			['Azure services with articles', '=COUNTIF(AzureServicesArticles[Article count],\">0\")',],
			['% Completion', '=B2/B1',],
		]

		for row in data:
			sheet.append(row)

		sheet['B3'].number_format = '0.0%'

		c = sheet.cell(row=1, column=4)
		c.value = 'Azure Service & Terraform Article Gap Analysis'
		c.font = Excel.styles.Font(size=16, bold=True)

		c = sheet.cell(row=2, column=4)
		c.value = 'This report shows the percentage of (Terraform-supported) Azure services for which a Terraform article is published.'
		
		c = sheet.cell(row=4, column=4)
		c.value = '\"Terraform-supported\" means having support via one or more Terraform resources in the azurerm provider.'

		c = sheet.cell(row=5, column=4)
		c.value = 'The other Terraform Azure providers (e.g., azuread) will be included in future reports.'

		# Write FAQ for purpose of report.
		c = sheet.cell(row=7, column=4)
		c.value = 'Questions that are answered via this report:'
		c.font = Excel.styles.Font(size=16, bold=True)

		c = sheet.cell(row=8, column=4)
		c.value = 'Q: What is the purpose of this report?'
		c = sheet.cell(row=9, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: To determine which Azure services do not have Terraform articles.\n'
		c.value += 'This information can help us to prioritize writing new Terraform samples and articles.'

		c = sheet.cell(row=11, column=4)
		c.value = 'Q: What percentage of Azure services have or don\'t have Terraform articles?'
		c = sheet.cell(row=12, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: See the \"% Completion\" value in cell B3.'

		c = sheet.cell(row=14, column=4)
		c.value = 'Q: Which Azure services have or don\'t have Terraform articles?'
		c = sheet.cell(row=15, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: The \"Azure services\" tab lists the Terraform article URLs for each Azure service.'

		c = sheet.cell(row=17, column=4)
		c.value = 'Q: How can I see which Terraform articles contain specific Terraform resources?'
		c = sheet.cell(row=18, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: The \"Terraform resources\" tab shows - for each Terraform resource - every article containing the resource name.\n'
		c.value += 'The \"Article (Y/N)\" column indicates whether the article is included in the report.\n'
		c.value += 'The \"Excluded articles\" tab lists the articles that are excluded from the report.'

		# Write FAQ for report's data.
		c = sheet.cell(row=20, column=4)
		c.value = 'Questions about the data:'
		c.font = Excel.styles.Font(size=16, bold=True)

		c = sheet.cell(row=21, column=4)
		c.value = 'Q: Where does the list of Azure services come from?'
		c = sheet.cell(row=22, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: Currently, from the azurerm reference documentation TOC on the HashiCorp site.\n'
		c.value += 'A future version of this report will retrieve the info from an API provided by the Terraform dev team.'

		c = sheet.cell(row=24, column=4)
		c.value = 'Q: What is the definition of a Terraform article?'
		c = sheet.cell(row=25, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: If an article contains a Terraform resource name and the article is not in the \"Excluded articles\" tab,\n'
		c.value += 'the article is counted as a Terraform article for the Azure service associated with the Terraform resource.'

		c = sheet.cell(row=27, column=4)
		c.value = 'Q: How do you determine if a Terraform resource is in an article?'
		c = sheet.cell(row=28, column=4)
		c.alignment = Alignment(wrapText=True)
		c.value = 'A: Currently, the report app searches  a static file built (6/25/2023) from the Bing search API.\n'
		c.value += 'The next version of this report will use data provided by the Terraform product group.'

	def _write_sheet_azure_services(self):
		'''Write Azure services worksheet sheet.'''

		# Create worksheet.
		sheet = self.wb.create_sheet(title='Azure services')

		# Write header.
		row_number = 1
		c = sheet.cell(row_number, column = 1)
		c.value = 'Azure service'
		c = sheet.cell(row_number, column = 2)
		c.value = 'Article count'
		c = sheet.cell(row_number, column = 3)
		c.value = f"Terraform (azurerm) articles for Azure service"

		# Set row number to 2 to skip header.
		row_number = 2

		# For each Azure service.
		for az_service in self.az_services:

			# Write Azure service name.
			c = sheet.cell(row_number, column = 1)
			c.value = f"{az_service.name}"

			# Write article count.
			c = sheet.cell(row_number, column = 2)
			c.alignment = Alignment(horizontal='center')
			c.value = len(az_service.articles)

			# Format found articles into single string with newline delimiter.
			found_articles = ''
			# For each article URL in the article list.
			for article_url in az_service.articles:
				# Add newline delimiter if not first article.
				if len(found_articles):
					found_articles += '\n' #'\015'
				# Add article URL to string.
				found_articles += f"{article_url}"

			# Write article URLs.
			c = sheet.cell(row_number, column = 3)
			c.alignment = Alignment(wrapText=True)
			c.value = f"{found_articles}"

			# Increment row number.
			row_number += 1

		# Create table.
		table = self._create_table(
			table_name='AzureServicesArticles', 
			row_number=row_number - 1, 
			column_letter='C')

		# Add table to sheet.
		sheet.add_table(table)

	def _create_table(self, table_name, row_number, column_letter):
		'''Create and return a table with the given name and row count.'''

		# Create table.
		table = Table(
			displayName=table_name, 
			ref=f"A1:{column_letter}{row_number}",
			 totalsRowShown=True)

		# Create a style with striped rows and banded columns
		style = TableStyleInfo(
			name="TableStyleMedium2", 
			showFirstColumn=False,
			showLastColumn=False, 
			showRowStripes=True, 
			showColumnStripes=False)

		# Apply table style.
		table.tableStyleInfo = style

		return table		

	def _write_sheet_terraform_resources(self):
		'''Write Terraform resources worksheet sheet.'''

		def _write_tf_resource_row(sheet, row_number, az_service, tf_resource_name, article_url=None):
			# Write Azure service name.
			c = sheet.cell(row_number, column = 1)
			c.value = f"{az_service.name}"

			# Write article count.
			c = sheet.cell(row_number, column = 2)
			c.value = f"{tf_resource_name}"

			# Write article URL.
			if article_url:
				c = sheet.cell(row_number, column = 3)
				c.value = f"{article_url}"

			# Write article (Y/N).
			c = sheet.cell(row_number, column = 4)
			if article_url and not az_service.is_article_excluded(article_url):
				c.value = f"Y"
			else:
				c.value = f"N"

		# Create worksheet.
		sheet = self.wb.create_sheet(title='Terraform resources')

		# Write header.
		row_number = 1
		c = sheet.cell(row_number, column = 1)
		c.value = 'Azure service'
		c = sheet.cell(row_number, column = 2)
		c.value = 'azurerm resource name'
		c = sheet.cell(row_number, column = 3)
		c.value = 'Article that contains the azurerm resource name'
		c = sheet.cell(row_number, column = 4)
		c.value = 'Article (Y/N)'

		row_number = 2
		for az_service in self.az_services:
			if len(az_service.search_results):
				for tf_resource_name, article_urls in az_service.search_results.items():
					if len(article_urls):
						for article_url in article_urls:
							_write_tf_resource_row(sheet, row_number, az_service, tf_resource_name, article_url)
							row_number += 1
					else:
						_write_tf_resource_row(sheet, row_number, az_service, tf_resource_name)
						row_number += 1
			else:
				_write_tf_resource_row(sheet, row_number, az_service, tf_resource_name)
				row_number += 1

		# Create table.
		table = self._create_table(
			table_name='TerraformServicesSearchResults', 
			row_number=row_number - 1, 
			column_letter='D')

		# Add table to sheet.
		sheet.add_table(table)

	def _write_sheet_excluded_articles(self, excluded_articles):
		# Write excluded articles to a new sheet.
		sheet = self.wb.create_sheet(title='Excluded articles')

		c = sheet.cell(row = 1, column = 1)
		c.value = f"Articles excluded from search results."
		
		for row_number, article_url in enumerate(excluded_articles, start=2):
			c = sheet.cell(row_number, column = 1)
			c.value = f"{article_url}*"